import requests
from bs4 import BeautifulSoup
import openpyxl
import chardet  # 导入chardet库

class download():
    # 创建一个函数来抓取页面内容并保存到文本文件
    def fetch_and_save_data(url, output_file):
        try:
            # 发送HTTP请求并获取页面内容
            response = requests.get(url)
            response.raise_for_status()

            # 使用chardet检测页面编码
            encoding = chardet.detect(response.content)['encoding']

            # 创建BeautifulSoup对象解析HTML，指定编码
            soup = BeautifulSoup(response.content, 'html.parser', from_encoding=encoding)

            # 提取<h1>标签内容
            h1_tag = soup.find('h1', {'class': 'title'})
            h1_text = h1_tag.text.strip() if h1_tag else ""

            # 提取<div>标签中的汉字
            content_div = soup.find('div', {'class': 'content', 'id': 'content'})
            content_text = ""
            if content_div:
                # 找到所有文本内容并连接起来
                content_parts = content_div.find_all(text=True)
                content_text = "\n".join([part.strip() for part in content_parts])

            # 写入提取的数据到txt文件
            with open(output_file, 'a', encoding='utf-8') as txt_file:
                txt_file.write(f"{h1_text}\n\n")
                txt_file.write(f"{content_text}\n\n")

            print(f"已成功抓取并保存数据到 {output_file}")
        except Exception as e:
            print(f"抓取并保存数据时出现错误: {e}")

if __name__ == '__main__':
    # 打开 Excel 目录文件
    workbook = openpyxl.load_workbook('content.xlsx')
    worksheet = workbook.active

    # 创建一个单独的txt文件来保存所有链接的内容，使用UTF-8编码
    all_data_file = "书籍名称可修改.txt"

    # 清空文件，以便重新写入
    with open(all_data_file, 'w', encoding='utf-8') as txt_file:
        txt_file.write("")  # 清空文件内容

    # 调用爬虫函数，从Excel文件中读取地址并抓取数据
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=2, max_col=2):
        url = row[0].value
        if url:
            download.fetch_and_save_data(url, all_data_file)

    # 关闭Excel文件
    workbook.close()

    print(f"已成功抓取并保存所有链接的内容到 {all_data_file}")
