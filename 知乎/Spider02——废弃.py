import requests
from bs4 import BeautifulSoup
import chardet  # 导入chardet库
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import openpyxl

class download():
    # 抓取封面
    def feng_mian(shuming,fm_url):
        try:
            chrome_option = Options()
            chrome_option.add_experimental_option('debuggerAddress', '127.0.0.1:9222')
            driver = webdriver.Chrome(executable_path='/Users/gaosg/chromedriver', options=chrome_option)

            # 打开网址
            driver.get(fm_url)
            driver.implicitly_wait(5)
            img_link = driver.find_element(By.XPATH, value='//*[@id="root"]/div/main/div/div/div[2]/div[2]/div/section/p/img').get_attribute('src')
            download.download_img(img_link, 'cover', shuming)
        except Exception as e:
            print(f"抓取并保存数据时出现错误: {e}")
    def download_img(url, local_path, file_name):
        reponse = requests.get(url)
        if reponse.status_code == 200:
            with open(f"{local_path}/{file_name}.jpg", 'wb') as f:
                f.write(reponse.content)
                print(f"图片已保存为{file_name}.jpg")
    # 创建一个函数来抓取页面内容并保存到文本文件
    def fetch_and_save_data(url, output_file):
        try:
            chrome_option = Options()
            chrome_option.add_experimental_option('debuggerAddress', '127.0.0.1:9222')
            driver = webdriver.Chrome(executable_path='/Users/gaosg/chromedriver', options=chrome_option)

            # 打开网址
            driver.get(url)
            driver.implicitly_wait(5)
            div = driver.find_element(By.XPATH,
                                           value='//*[@id="root"]/div/main/div/div/div[2]/div[2]/div/section/div').get_attribute(
                'outerHTML')
            #print(div)
            # 创建BeautifulSoup对象解析HTML，指定编码
            soup = BeautifulSoup(div, 'html.parser')
            # 提取<p>标签中的汉字
            with open(output_file, 'a', encoding='utf-8') as txt_file:
                h1_tag = soup.find('h1')
                h1_text = h1_tag.text.strip() if h1_tag else ""
                txt_file.write(f"{h1_text}\n")
                for content_p in soup.find_all('p'):
                    content_parts = content_p.find_all(text=True)
                    content_text = "".join([part.strip() for part in content_parts])
                    txt_file.write(f"{content_text}\n")

            print(f"已成功抓取并保存数据到 {output_file}")
        except Exception as e:
            print(f"抓取并保存数据时出现错误: {e}")

if __name__ == '__main__':
    shuming = '法医秦明'

    # 打开 Excel 目录文件
    workbook = openpyxl.load_workbook('shu.xlsx')
    worksheet = workbook.active

    # 创建一个单独的txt文件来保存所有链接的内容，使用UTF-8编码
    all_data_file = "cover/"+shuming+".txt"

    # 清空文件，以便重新写入
    with open(all_data_file, 'w', encoding='utf-8') as txt_file:
        txt_file.write("")

    # 调用爬虫函数，从Excel文件中读取封面地址并抓取图片
    fm_link = worksheet.cell(1, 2).value
    if fm_link:
        download.feng_mian(shuming, fm_link)

    # 调用爬虫函数，从Excel文件中读取地址并抓取内容
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=2, max_col=2):
        url = row[0].value
        if url:
            download.fetch_and_save_data(url, all_data_file)

    # 关闭Excel文件
    workbook.close()
