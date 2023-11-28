import os

import requests
from bs4 import BeautifulSoup
from ebooklib import epub
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import openpyxl

class download():
    # 抓取封面
    def feng_mian(shuming,fm_url):
        try:
            chrome_option = Options()
            chrome_option.add_experimental_option('debuggerAddress', '127.0.0.1:9222')
            driver = webdriver.Chrome(executable_path='/Users/gaosg/chromedriver', options=chrome_option)

            # 打开网址
            driver.get(fm_url)
            driver.implicitly_wait(5)
            try:
                img_link = driver.find_element(By.XPATH, value='//*[@id="root"]/div/main/div/div/div[2]/div[2]/div/section/p/img').get_attribute('src')
            except:
                img_link = driver.find_element(By.XPATH,value='//*[@id="root"]/div/main/div/div/div[2]/div[2]/div/section/div/p/img').get_attribute('src')
            download.download_img(img_link, 'cover', shuming)
        except Exception as e:
            print(f"抓取并保存数据时出现错误: {e}")
    def download_img(url, local_path, file_name):
        reponse = requests.get(url)
        if reponse.status_code == 200:
            with open(f"{local_path}/{file_name}.jpg", 'wb') as f:
                f.write(reponse.content)
                print(f"图片已保存为{file_name}.jpg")

    #返回书籍章节内容
    def re_cont(url):
        try:
            chrome_option = Options()
            chrome_option.add_experimental_option('debuggerAddress', '127.0.0.1:9222')
            driver = webdriver.Chrome(executable_path='/Users/gaosg/chromedriver', options=chrome_option)

            # 打开网址
            driver.get(url)
            driver.implicitly_wait(60)
            div = driver.find_element(By.XPATH,
                                           value='//*[@id="root"]/div/main/div/div/div[2]/div[2]/div/section/div').get_attribute(
                'outerHTML')
            #print(div)
            # 创建BeautifulSoup对象解析HTML，指定编码
            soup = BeautifulSoup(div, 'html.parser')
            # 提取<p>标签中的汉字
            content_text = []
            for content_p in soup.find_all('p'):
                content_parts = content_p.find_all(text=True)
                content_text.append("".join([part.strip() for part in content_parts]))
            content_v = '<br>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;'.join(content_text)
            #cont = soup
            cont = content_v
            return cont
        except Exception as e:
            print(f"抓取并保存数据时出现错误: {e}")

    def check_img(img_path):
        if os.path.exists(img_path):
            return True
        else:
            return False


if __name__ == '__main__':
    shuming = '侦探推理游戏2'
    zuozhe = '罗非鱼' #佚名
    # 打开 Excel 目录文件
    workbook = openpyxl.load_workbook(f'cover/{shuming}.xlsx')
    worksheet = workbook.active

    # 调用爬虫函数，从Excel文件中读取封面地址并抓取图片
    fm_link = worksheet.cell(1, 2).value
    if fm_link:
        download.feng_mian(shuming, fm_link)

    # 创建一个EpubBook对象
    book = epub.EpubBook()

    # 设置书籍的元数据
    book.set_identifier('id123456')
    book.set_title(shuming)
    book.set_language('zh')
    book.add_author(zuozhe)

    #调用返回书籍章节名称和章节内容
    chapters = []

    path_image = f'cover/{shuming}.jpg'
    fm_image = f'cover/fengmian.jpg'
    cover_image = download.check_img(path_image)
    if cover_image == True:
        min_row_value = 2
    else:
        min_row_value = 1
    for row in worksheet.iter_rows(min_row=min_row_value, max_row=worksheet.max_row, min_col=1, max_col=2):
        title_v = row[0].value
        url = row[1].value
        cont = download.re_cont(url)
        chapter_info = {'h1':title_v, 'cont':cont}
        chapters.append(chapter_info)
    # 关闭Excel文件
    workbook.close()

    a = 0
    b = []
    for chapter in chapters:
        chapter_title = chapter['h1']
        chapter_cont = chapter['cont']
        #添加章节对象

        a = a+1
        chapter_file = epub.EpubHtml(title=chapter_title, file_name=f'chap_{a}.xhtml', lang='zh')
        chapter_file.content = f'<h2>{chapter_title}</h2><p>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;{chapter_cont}</p>'
        #chapter_file.content = f'<p>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;{chapter_cont}</p>'
        # 将章节添加到书籍中
        book.add_item(chapter_file)
        # 定义书籍的目录结构
        book.toc.append(chapter_file)
        # 定义书籍的Spine（内容顺序）
        b.append(chapter_file)
        book.spine = b

    # 设置封面
    if cover_image == True:
        book.set_cover("cover.jpg", open(path_image, 'rb').read())
    else:
        book.set_cover("cover.jpg", open(fm_image, 'rb').read())

    # 添加默认的NCX和NAV文件
    book.add_item(epub.EpubNcx())
    book.add_item(epub.EpubNav())
    # 生成epub文件
    epub.write_epub(f"{shuming}.epub", book)
