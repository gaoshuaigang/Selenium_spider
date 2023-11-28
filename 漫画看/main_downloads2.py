# -*- coding: utf-8 -*-
# 先启动自定义浏览器
# /Applications/Google\ Chrome.app/Contents/MacOS/Google\ Chrome --remote-debugging-port=9222
import os

import openpyxl
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By

class download():
    def start_download(self):
        chrome_option = Options()
        chrome_option.add_argument('headless')
        chrome_option.add_argument('no-sandbox')
        chrome_option.add_argument('ignore-certificate-errors')
        chrome_option.add_argument('disable-gpu')
        #chrome_option.add_experimental_option('debuggerAddress', '127.0.0.1:9222')
        driver = webdriver.Chrome(executable_path='/Users/gaosg/chromedriver', options=chrome_option)

        wb = openpyxl.load_workbook('People/output_1.xlsx')
        sheet = wb['451-500']
        num = []
        for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if i == 1:  # 跳过第一行
                continue
            num.append(list(row))
        #print(num)

        # 创建保存Excel文件的文件夹
        folder_path = 'People'
        os.makedirs(folder_path, exist_ok=True)

        # 检查是否存在已有的Excel文件
        try:
            workbook = openpyxl.load_workbook('People/An_people.xlsx')
            sheet = workbook.active
        except FileNotFoundError:
            # 如果文件不存在，则创建一个新的Excel文件
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            # 添加标题行
            sheet.append(['章节', '地址链接'])

        #取每一行的值
        for row in num:
            cell = row
            #行值中的第一个值、第二个值、第三个值，忽略excel中的序号列
            title = cell[1]
            page = cell[3]
            url = cell[2]
            #对页码进行便利，页码数17则从1查询到17页
            for p in range(1,int(page)+1):
                #输出标题定义为titlepage，输出图片地址为src
                titlepage = str(title)+'page'+str(p)
                #打开网页
                driver.get(str(url)+'?p=' + str(p))
                driver.implicitly_wait(5)
                #获取图片所在位置的HTML源码
                html1 = driver.find_element(By.ID, 'images').get_attribute('outerHTML')
                #解析返回的html内容
                soup = BeautifulSoup(html1, 'html.parser')
                img = soup.find('img')
                #筛选img中的src属性
                src = img['src']
                #print(titlepage, src)
                # 将span内容和href写入Excel表格的新行
                sheet.append([titlepage, src])

                # 保存Excel文件
                file_path = os.path.join(folder_path, 'An_people.xlsx')
                workbook.save(file_path)
            print(title)

if __name__ == '__main__':
    download.start_download(webdriver)
